import os
import io
import re
import pandas as pd
import pdfplumber
import pypdf
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def detect_bank(pdf_path):
    """
    Detects the bank/layout from the statement text or filename.
    Returns one of: 'RBC', 'TD', 'BMO', 'CIBC', 'Tangerine', 'Vancity', or 'Standard'
    """
    text = ""
    try:
        if hasattr(pdf_path, "seek"):
            pdf_path.seek(0)
            reader = pypdf.PdfReader(pdf_path)
            for page in reader.pages[:2]: # Check first 2 pages
                t = page.extract_text()
                if t:
                    text += t
        else:
            with open(pdf_path, "rb") as f:
                reader = pypdf.PdfReader(f)
                for page in reader.pages[:2]:
                    t = page.extract_text()
                    if t:
                        text += t
    except Exception:
        pass
        
    text_lower = text.lower()
    
    # Auto-detect using keywords
    if "royal bank" in text_lower or "rbc" in text_lower:
        return "RBC"
    if "td canada trust" in text_lower or "td bank" in text_lower:
        return "TD"
    if "bank of montreal" in text_lower or "bmo" in text_lower:
        return "BMO"
    if "cibc" in text_lower:
        return "CIBC"
    if "tangerine" in text_lower:
        return "Tangerine"
    if "vancity" in text_lower or "vancouver city savings" in text_lower:
        return "Vancity"
        
    # Check filename as fallback
    filename = ""
    if isinstance(pdf_path, str):
        filename = os.path.basename(pdf_path).lower()
    elif hasattr(pdf_path, "name"):
        filename = str(pdf_path.name).lower()
        
    if "rbc" in filename: return "RBC"
    if "td" in filename: return "TD"
    if "bmo" in filename: return "BMO"
    if "cibc" in filename: return "CIBC"
    if "tangerine" in filename: return "Tangerine"
    if "vancity" in filename: return "Vancity"
    
    return "Standard"

def extract_statement_period(text):
    """
    Extracts (start_year, start_month, end_year, end_month) from statement text.
    Returns integers. Default to current year if not found.
    """
    months_map = {
        "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
        "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12
    }
    
    current_year = datetime.now().year
    
    # Try to find dates like "December 20, 2024 to January 19, 2025" or "Dec 20, 2024 - Jan 19, 2025"
    pattern = r'\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\s+\d{1,2},\s+(\d{4})\b'
    matches = re.findall(pattern, text.lower())
    
    # Parse unique dates and sort them chronologically
    parsed_dates = sorted(list(set((int(y), months_map[m]) for m, y in matches)))
    if len(parsed_dates) >= 2:
        y1, m1 = parsed_dates[0]
        y2, m2 = parsed_dates[-1]
        return y1, m1, y2, m2
    elif len(parsed_dates) == 1:
        y1, m1 = parsed_dates[0]
        return y1, m1, y1, m1
        
    # Fallback to general years
    year_pattern = re.compile(r'\b(?:19|20)\d{2}\b')
    years = [int(y) for y in year_pattern.findall(text)]
    if len(years) >= 2:
        return years[0], 1, years[1], 12
    elif len(years) == 1:
        return years[0], 1, years[0], 12
        
    return current_year, 1, current_year, 12

def extract_statement_year_range(text):
    """
    Deprecated: Use extract_statement_period instead.
    """
    y1, m1, y2, m2 = extract_statement_period(text)
    return y1, y2

def parse_date(date_str, start_year, start_month, end_year, end_month):
    """
    Attempts to parse date strings like 'Jan 1', 'Jan 01', '01/02', '1/2'.
    Uses mathematical transition logic based on statement period start/end months.
    """
    months_map = {
        "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
        "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12
    }
    
    clean_str = str(date_str).strip().lower().replace(",", "")
    
    if clean_str.isdigit():
        return None, None
    
    # Check if a 4-digit year is in the string
    year_match = re.search(r'\b(20\d{2}|19\d{2})\b', clean_str)
    year = None
    if year_match:
        year = int(year_match.group(1))
        clean_str = re.sub(r'\b(20\d{2}|19\d{2})\b', '', clean_str).strip()
        
    # Check for Month Day format (e.g., "Jan 15", "Dec 2")
    m = re.match(r'^([a-z]{3})\s+(\d{1,2})$', clean_str)
    if m:
        month_name, day_val = m.groups()
        month_num = months_map.get(month_name, 1)
        day_num = int(day_val)
        
        if year is None:
            if start_month <= end_month:
                year = start_year
            else:
                if month_num >= start_month:
                    year = start_year
                else:
                    year = end_year
                    
        return f"{year}-{month_num:02d}-{day_num:02d}", month_num
        
    # Check for numerical Slash format (e.g., "12/25", "12-25", "01/02")
    m = re.match(r'^(\d{1,2})[/\-](\d{1,2})$', clean_str)
    if m:
        val1, val2 = m.groups()
        num1, num2 = int(val1), int(val2)
        if num1 > 12: # must be DD/MM
            month_num, day_num = num2, num1
        else: # assume MM/DD
            month_num, day_num = num1, num2
            
        if year is None:
            if start_month <= end_month:
                year = start_year
            else:
                if month_num >= start_month:
                    year = start_year
                else:
                    year = end_year
                    
        return f"{year}-{month_num:02d}-{day_num:02d}", month_num
        
    # Standard YYYY-MM-DD or MM/DD/YYYY fallback (respect statement year)
    try:
        dt = pd.to_datetime(clean_str)
        month_num = dt.month
        if year is None:
            if start_month <= end_month:
                year = start_year
            else:
                if month_num >= start_month:
                    year = start_year
                else:
                    year = end_year
        return f"{year}-{month_num:02d}-{dt.day:02d}", month_num
    except Exception:
        pass
        
    return None, None

def looks_like_date_word(word):
    """
    Checks if a string word is a valid date token (digits, month abbreviation, month name).
    """
    w = str(word).lower().strip().strip(",.*:()#")
    if not w:
        return False
    if w.isdigit():
        if len(w) == 4 and (w.startswith("19") or w.startswith("20")):
            return False
        return True
    months = {
        "jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec",
        "january", "february", "march", "april", "june", "july", "august", "september", "october", "november", "december",
        "janv", "fevr", "mars", "avr", "mai", "juin", "juil", "aout", "sept", "dece", "decembre"
    }
    if w in months:
        return True
    if "/" in w or "-" in w:
        if any(c.isdigit() for c in w):
            return True
    return False

def is_disclaimer_or_metadata(desc_text):
    """
    Checks if a description string belongs to footer metadata,
    legal disclaimers, or statement headers.
    """
    txt = str(desc_text).lower().strip()
    if not txt:
        return False
        
    patterns = [
        r'\bpage \d+',
        r'^page\b',
        r'\bpage of\b',
        r'continued on next page',
        r'trademark of',
        r'registered trademark',
        r'interac is a registered',
        r'important:',
        r'foreign currency conversion',
        r'foreign currency exchange',
        r'cibc account statement',
        r'account summary',
        r'branch transit number',
        r'opening balance on',
        r'closing balance on',
        r'statement period',
        r'for questions on this update',
        r'contact us by phone',
        r'tty hearing impaired',
        r'outside canada',
        r'www\.cibc\.com',
        r'balance forward',
        r'transaction details',
        r'\bper-20\d{2}\b',
        r'^\d{4,}\s+per-\d+$',
        r'bankbook or paperless',
        r'statement: \d+ days',
        r'this rule does not apply',
        r'your rights under your',
        r'if you withdraw foreign',
        r'\bof cibc\b',
        r'transaction amount reflects',
        r'exchange rate displayed',
        r'converted amount',
        r'converted to canadian',
        r'information about your',
        r'how we charge interest',
        r'grace period',
        r'installment plan',
        r'convert an eligible',
        r'payment period extensions',
        r'minimum payment',
        r'refer to the cibc',
        r'denotes transaction in',
        r'your interest',
        r'total interest',
        r'annual interest rate',
        r'your payments',
        r'payments and credits',
        r'pre-authorized payment',
        r'amount due',
        r'charges and credits',
        r'card number',
        r'\btrans\b',
        r'\bpost\b',
        r'\bdescription\b',
        r'\bspend categories\b',
        r'^your$',
        r'^interest$',
        r'^payments$',
        r'^charges$',
        r'^credits$',
        r'^period$',
        r'this period'
    ]
    for p in patterns:
        if re.search(p, txt):
            return True
    return False

def extract_digital_pdf(pdf_path, bank_name):
    """
    Extracts transactions from a digital text statement PDF using pdfplumber coordinates.
    Filters columns cleanly and groups multiline descriptions.
    """
    raw_text = ""
    try:
        if hasattr(pdf_path, "seek"):
            pdf_path.seek(0)
            reader = pypdf.PdfReader(pdf_path)
            for page in reader.pages:
                t = page.extract_text()
                if t: raw_text += t
        else:
            with open(pdf_path, "rb") as f:
                reader = pypdf.PdfReader(f)
                for page in reader.pages:
                    t = page.extract_text()
                    if t: raw_text += t
    except Exception:
        pass
        
    # 1. Determine the statement period dates context
    start_year, start_month, end_year, end_month = extract_statement_period(raw_text)
    is_credit_card = "visa" in raw_text.lower() or "mastercard" in raw_text.lower() or "card" in raw_text.lower()
    
    prev_bal = None
    ending_bal = None
    if is_credit_card:
        txt_lower = raw_text.lower()
        # Find previous balance
        m_prev = re.search(r'previous\s+balance\s+[\-\$]*\s*([\d,]+\.\d{2})', txt_lower)
        if m_prev:
            try:
                prev_bal = float(m_prev.group(1).replace(",", ""))
            except ValueError:
                pass
        # Find ending balance
        m_end = re.search(r'(?:total|new|ending|closing)\s+balance\s*(?:=|\s)\s*[\-\$]*\s*([\d,]+\.\d{2})', txt_lower)
        if m_end:
            try:
                ending_bal = float(m_end.group(1).replace(",", ""))
            except ValueError:
                pass
    
    raw_rows = []
    
    # 2. Extract tables/lines using pdfplumber coordinates for column positions
    try:
        if hasattr(pdf_path, "seek"):
            pdf_path.seek(0)
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                words = page.extract_words()
                # Find vertical coordinates of column headers inside the actual Transaction table header row
                header_top = None
                post_x0 = None
                desc_x0 = None
                for w in words:
                    w_text = w["text"].lower()
                    if w_text in ("date", "description"):
                        if header_top is None:
                            header_top = w["top"]
                    if w_text == "description":
                        if desc_x0 is None:
                            desc_x0 = w["x0"]
                    if w_text == "post":
                        if post_x0 is None:
                            post_x0 = w["x0"]
                
                # Check column X coordinates
                debit_x_coords = []
                credit_x_coords = []
                balance_x_coords = []
                
                if header_top is not None:
                    for w in words:
                        # Allow height tolerance of 5 pt to capture all headers on the same row
                        if abs(w["top"] - header_top) <= 5.0:
                            w_text = w["text"].lower()
                            if w_text in ("withdrawals", "debit", "payments", "charges", "withdrawals($)"):
                                debit_x_coords.append((w["x0"], w["x1"]))
                            elif w_text in ("deposits", "credit", "receipts", "deposits($)"):
                                credit_x_coords.append((w["x0"], w["x1"]))
                            elif w_text in ("balance", "balance($)"):
                                balance_x_coords.append((w["x0"], w["x1"]))
                
                # Default fallback X ranges based on bank layouts
                if is_credit_card:
                    deb_range = (999.0, 999.0)
                    cred_range = (999.0, 999.0)
                    bal_range = (500.0, 580.0)
                elif bank_name == "RBC":
                    deb_range = (300.0, 410.0)
                    cred_range = (410.0, 510.0)
                    bal_range = (510.0, 600.0)
                elif bank_name == "TD":
                    deb_range = (280.0, 380.0)
                    cred_range = (380.0, 480.0)
                    bal_range = (480.0, 600.0)
                elif bank_name == "CIBC":
                    deb_range = (320.0, 430.0)
                    cred_range = (430.0, 520.0)
                    bal_range = (520.0, 600.0)
                else:
                    # Standard fallback ranges
                    deb_range = (300.0, 400.0)
                    cred_range = (400.0, 500.0)
                    bal_range = (500.0, 600.0)
                
                # Update ranges if actual table header coordinates detected (with safety limit > 200)
                if debit_x_coords:
                    min_x = min(coord[0] for coord in debit_x_coords)
                    max_x = max(coord[1] for coord in debit_x_coords)
                    if min_x > 200.0:
                        deb_range = (min_x - 10, max_x + 10)
                if credit_x_coords:
                    min_x = min(coord[0] for coord in credit_x_coords)
                    max_x = max(coord[1] for coord in credit_x_coords)
                    if min_x > 200.0:
                        cred_range = (min_x - 10, max_x + 10)
                if balance_x_coords:
                    min_x = min(coord[0] for coord in balance_x_coords)
                    max_x = max(coord[1] for coord in balance_x_coords)
                    if min_x > 200.0:
                        bal_range = (min_x - 10, max_x + 10)
                
                # Group words into lines based on vertical coordinate (top)
                # Group words by rounded top value (to group cells on same horizontal row)
                lines_dict = {}
                for w in words:
                    top = round(w["top"], 1)
                    found = False
                    for existing_top in lines_dict.keys():
                        if abs(existing_top - w["top"]) <= 6.0: # line height tolerance
                            lines_dict[existing_top].append(w)
                            found = True
                            break
                    if not found:
                        lines_dict[w["top"]] = [w]
                
                # Sort lines vertically from top to bottom
                active_extraction = False
                for top_val in sorted(lines_dict.keys()):
                    line_words = sorted(lines_dict[top_val], key=lambda x: x["x0"])
                    line_txt = " ".join([w["text"] for w in line_words]).strip()
                    line_txt_lower = line_txt.lower()
                    
                    # 1. Detect start and end markers
                    if is_credit_card:
                        if "your payments" in line_txt_lower or "your new charges and credits" in line_txt_lower:
                            active_extraction = True
                            continue
                        
                        # Stop if footer or summary categories encountered
                        if any(marker in line_txt_lower for marker in [
                            "information about your cibc",
                            "cibc creditsmart spend report",
                            "your message centre",
                            "go paperless",
                            "total for 4500"
                        ]):
                            active_extraction = False
                            continue
                            
                    else:
                        active_extraction = True # Default to true for standard bank statements
                        
                    if not active_extraction:
                        continue
                        
                    if header_top is not None and top_val < (header_top - 2.0):
                        continue
                        
                    # Skip total and summary lines early before coordinate filtering
                    if "total for" in line_txt_lower or "total interest" in line_txt_lower or "total payments" in line_txt_lower or line_txt_lower.startswith("total ") or line_txt_lower == "total":
                        continue
                    
                    # Group words into tokens based on X coordinates
                    # Check X coordinates of numeric tokens
                    date_tokens = []
                    desc_tokens = []
                    debit_tokens = []
                    credit_tokens = []
                    balance_tokens = []
                    
                    for w in line_words:
                        x_mid = (w["x0"] + w["x1"]) / 2.0
                        text_token = w["text"]
                        
                        if text_token.strip() == "Q":
                            continue
                        
                        # Clean amounts helper
                        is_numeric = re.match(r'^\-?\$?\d+[\d,\.]*$', text_token)
                        
                        date_limit = (post_x0 - 2.0) if (is_credit_card and post_x0 is not None) else (70.0 if is_credit_card else 95.0)
                        desc_limit = (desc_x0 - 5.0) if (is_credit_card and desc_x0 is not None) else (110.0 if is_credit_card else 95.0)
                        
                        if is_credit_card and date_limit <= x_mid < desc_limit:
                            continue
                            
                        if x_mid < date_limit and looks_like_date_word(text_token): # Leftmost is date (e.g. 'Jan 01')
                            date_tokens.append(text_token)
                        elif deb_range[0] <= x_mid < deb_range[1] and is_numeric:
                            debit_tokens.append(text_token)
                        elif cred_range[0] <= x_mid < cred_range[1] and is_numeric:
                            credit_tokens.append(text_token)
                        elif bal_range[0] <= x_mid < bal_range[1] and is_numeric:
                            balance_tokens.append(text_token)
                        else:
                            desc_tokens.append(text_token)
                            
                    date_str = " ".join(date_tokens).strip()
                    desc_str = " ".join(desc_tokens).strip()
                    debit_str = "".join(debit_tokens).replace("$", "").replace(",", "").strip()
                    credit_str = "".join(credit_tokens).replace("$", "").replace(",", "").strip()
                    balance_str = "".join(balance_tokens).replace("$", "").replace(",", "").strip()
                    
                    # Enforce cleaning of OCR description line-by-line
                    if is_disclaimer_or_metadata(desc_str):
                        continue
                        
                    # Filter out empty spacer rows
                    if date_str or desc_str or debit_str or credit_str or balance_str:
                        raw_rows.append({
                            "date_raw": date_str,
                            "description": desc_str,
                            "debit_raw": debit_str,
                            "credit_raw": credit_str,
                            "balance_raw": balance_str
                        })
    except Exception as e:
        print(f"Error extracting digital pdf text via coordinates: {e}")
        return [], 0.0
        
    # 3. Process, merge multiline descriptions, and parse values
    transactions = []
    prev_date = None
    prev_month_num = None
    
    # Store opening balance
    opening_bal = prev_bal if (is_credit_card and prev_bal is not None) else 0.0
    opening_found = True if (is_credit_card and prev_bal is not None) else False
    seen_closing_balance = False
    
    for r in raw_rows:
        date_raw = r["date_raw"]
        desc = r["description"]
        
        if seen_closing_balance:
            continue
            
        desc_lower = desc.lower()
        if "closing balance" in desc_lower or "ending balance" in desc_lower:
            seen_closing_balance = True
            continue
            
        if "date" in desc_lower and "description" in desc_lower:
            continue
            
        # Skip disclaimer/metadata rows
        if is_disclaimer_or_metadata(desc):
            continue
            
        desc_clean = desc.lower().strip()
        if desc_clean.startswith("total") or desc_clean in ("subtotal", "subtotals", "page total", "closing balance"):
            continue
            
        debit_raw = r["debit_raw"]
        credit_raw = r["credit_raw"]
        balance_raw = r["balance_raw"]
        
        # Check if this row is opening balance
        desc_lower = desc.lower()
        if "opening" in desc_lower or "balance forward" in desc_lower:
            try:
                # Try to parse balance
                val_bal = float(balance_raw) if balance_raw else (float(debit_raw) if debit_raw else 0.0)
                opening_bal = val_bal
                opening_found = True
            except ValueError:
                pass
            continue
            
        # If this is a multiline description continuation (no date, no amount, no balance)
        if not date_raw and not debit_raw and not credit_raw and not balance_raw:
            if transactions and desc:
                # Append description to previous transaction
                transactions[-1]["description"] += " " + desc
            continue
            
        # Check if this row is a continuation row containing amounts for the previous transaction
        temp_debit = None
        temp_credit = None
        temp_balance = None
        try:
            if debit_raw: temp_debit = float(debit_raw)
        except ValueError: pass
        try:
            if credit_raw: temp_credit = float(credit_raw)
        except ValueError: pass
        try:
            if balance_raw: temp_balance = float(balance_raw)
        except ValueError: pass

        if not date_raw and (temp_debit is not None or temp_credit is not None or temp_balance is not None):
            if transactions and transactions[-1]["debit"] is None and transactions[-1]["credit"] is None and transactions[-1]["balance"] is None:
                if desc:
                    transactions[-1]["description"] += " " + desc
                transactions[-1]["debit"] = temp_debit
                transactions[-1]["credit"] = temp_credit
                transactions[-1]["balance"] = temp_balance
                continue

        # Parse numeric values
        debit = None
        credit = None
        balance = None
        
        try:
            if debit_raw: debit = float(debit_raw)
        except ValueError: pass
        
        try:
            if credit_raw: credit = float(credit_raw)
        except ValueError: pass
        
        try:
            if balance_raw: balance = float(balance_raw)
        except ValueError: pass
        
        # If credit card statement, amount is in balance, determine debit/credit
        if is_credit_card and balance is not None:
            val_amt = balance
            balance = None
            desc_lower = desc.lower()
            if "payment" in desc_lower or "thank you" in desc_lower or "paiment" in desc_lower or "refund" in desc_lower or "credit" in desc_lower or "rebate" in desc_lower or val_amt < 0:
                credit = abs(val_amt)
                debit = None
            else:
                debit = val_amt
                credit = None
        
        # Parse date
        date_parsed = None
        if date_raw:
            date_parsed, month_num = parse_date(date_raw, start_year, start_month, end_year, end_month)
            if date_parsed:
                prev_date = date_parsed
                prev_month_num = month_num
        else:
            date_parsed = prev_date
            
        # Clean description whitespace and remove repeated spaces
        desc_cleaned = re.sub(r'\s+', ' ', str(desc)).strip()
        
        # Eliminate rows with OCR artifacts trailing or metadata fragments
        if is_disclaimer_or_metadata(desc_cleaned):
            continue
            
        # Remove trailing disclaimer lines that bleed into descriptions
        disclaimer_phrases = [
            "about your cibc", "aventura visa", "minimum payment", "pre-authorized payment",
            "regular purchases", "cash advances", "interest rates", "annual rate",
            "identifies points multiplier", "foreign currency", "convenience cheques",
            "if you find an error", "how we charge", "grace period", "installment plan",
            "payment options", "go paperless", "important notice", "message centre"
        ]
        for phrase in disclaimer_phrases:
            if phrase in desc_cleaned.lower():
                idx_phrase = desc_cleaned.lower().find(phrase)
                desc_cleaned = desc_cleaned[:idx_phrase].strip()
                
        # Clean final trailing characters
        desc_cleaned = re.sub(r'\s+', ' ', desc_cleaned).strip()
        
        # Strict validation: transaction must have date, description, and at least one amount
        if date_parsed and desc_cleaned and (debit is not None or credit is not None or balance is not None):
            transactions.append({
                "date": date_parsed,
                "description": desc_cleaned,
                "debit": debit,
                "credit": credit,
                "balance": balance,
                "is_credit_card": is_credit_card
            })
            
    # Remove any completely empty or invalid transactions
    final_txs = []
    for tx in transactions:
        if not tx["date"] or not tx["description"]:
            continue
        if tx["debit"] is None and tx["credit"] is None and tx["balance"] is None:
            continue
        if is_credit_card and ending_bal is not None:
            tx["statement_ending_balance"] = ending_bal
        final_txs.append(tx)
        
    return final_txs, opening_bal

def reconcile_transactions(transactions, opening_balance):
    """
    Validates balance flows and calculates metrics.
    Opening Balance + sum(credits) - sum(debits) = Closing Balance
    """
    total_debits = 0.0
    total_credits = 0.0
    
    for tx in transactions:
        total_debits += tx.get("debit") if tx.get("debit") else 0.0
        total_credits += tx.get("credit") if tx.get("credit") else 0.0
        
    is_cc = any(tx.get("is_credit_card") for tx in transactions)
    # Calculate closing balance from transactions
    if is_cc:
        calculated_closing = opening_balance - total_credits + total_debits
    else:
        calculated_closing = opening_balance - total_debits + total_credits
    
    # Actual closing balance from last transaction if available
    actual_closing = opening_balance
    
    # Try to find statement ending balance metadata first
    stmt_ending = next((tx.get("statement_ending_balance") for tx in transactions if tx.get("statement_ending_balance") is not None), None)
    if stmt_ending is not None:
        actual_closing = stmt_ending
    elif transactions:
        # Find the last transaction with a valid balance
        for tx in reversed(transactions):
            if tx.get("balance") is not None:
                actual_closing = tx.get("balance")
                break
                
    diff = round(actual_closing - calculated_closing, 2)
    reconciled = abs(diff) <= 0.05
    
    return {
        "opening_balance": opening_balance,
        "closing_balance": actual_closing,
        "total_withdrawals": total_debits,
        "total_deposits": total_credits,
        "transaction_count": len(transactions),
        "difference": diff,
        "reconciled": reconciled,
        "warning": "" if reconciled else f"Reconciliation Warning: Balance mismatch of ${diff:,.2f}"
    }

def apply_excel_category_map(df, mapping_excel_path):
    """
    Reads dynamic keyword mapping files from an uploaded Excel sheet.
    Excel columns required: Keyword, Category Name, Excel Column Name, GST Rate, PST Rate, GIFI Code.
    """
    try:
        map_df = pd.read_excel(mapping_excel_path)
        # Standardize column headers to lowercase
        map_df.columns = [str(col).strip().lower() for col in map_df.columns]
        
        # Compile list of rules
        mappings = []
        for _, row in map_df.iterrows():
            kw = str(row.get("keyword", "")).strip().lower()
            cat = str(row.get("category name", row.get("category", ""))).strip()
            gifi = str(row.get("gifi code", row.get("gifi", ""))).strip()
            gst = str(row.get("gst rate", row.get("gst", ""))).strip()
            
            if kw and cat:
                mappings.append({
                    "keyword": kw,
                    "category": cat,
                    "gifi": gifi,
                    "gst": gst
                })
                
        # Apply rules to df
        for idx, row in df.iterrows():
            desc_norm = str(row["description"]).strip().lower()
            
            for rule in mappings:
                if rule["keyword"] in desc_norm:
                    df.at[idx, "category"] = rule["category"]
                    if rule["gifi"]:
                        df.at[idx, "gifi_code"] = rule["gifi"]
                    if rule["gst"]:
                        df.at[idx, "gst_rate"] = rule["gst"]
                    break # prioritize first matching keyword rule
    except Exception as e:
        print(f"Error applying Excel category map: {e}")
        
    return df

def generate_excel_report(transactions, reconciliation, bank_name, filename):
    """
    Builds a beautifully styled, premium multi-sheet Excel workbook.
    """
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    
    # 1. Sheet 1: Summary Dashboard
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Title styling
    title_font = Font(name="Calibri", size=16, bold=True, color="1F497D")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    accent_fill = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    double_bottom_border = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )
    
    # Set up Summary Dashboard
    ws_summary.append(["BANK STATEMENT SUMMARY REPORT"])
    ws_summary.cell(1, 1).font = title_font
    ws_summary.row_dimensions[1].height = 30
    ws_summary.append([]) # spacer
    
    metrics = [
        ("Parameter", "Value"),
        ("Institution / Bank", bank_name),
        ("Source File", filename),
        ("Transaction Count", reconciliation["transaction_count"]),
        ("Opening Balance", reconciliation["opening_balance"]),
        ("Total Withdrawals (Debits)", -reconciliation["total_withdrawals"]),
        ("Total Deposits (Credits)", reconciliation["total_deposits"]),
        ("Closing Balance", reconciliation["closing_balance"]),
        ("Reconciliation Difference", reconciliation["difference"]),
        ("Reconciliation Status", "RECONCILED" if reconciliation["reconciled"] else "MISMATCH WARNING")
    ]
    
    for label, val in metrics:
        ws_summary.append([label, val])
        
    # Style Summary Grid
    ws_summary.merge_cells("A1:B1")
    ws_summary.row_dimensions[3].height = 20
    ws_summary.cell(3, 1).font = header_font
    ws_summary.cell(3, 1).fill = header_fill
    ws_summary.cell(3, 2).font = header_font
    ws_summary.cell(3, 2).fill = header_fill
    
    for r_idx in range(4, 14):
        ws_summary.cell(r_idx, 1).font = bold_font
        ws_summary.cell(r_idx, 1).fill = accent_fill
        ws_summary.cell(r_idx, 1).border = thin_border
        
        val_cell = ws_summary.cell(r_idx, 2)
        val_cell.font = normal_font
        val_cell.border = thin_border
        
        # Number formats
        if r_idx in (7, 8, 9, 10, 11):
            val_cell.number_format = "$#,##0.00"
            
        if r_idx == 12: # status warning coloring
            if reconciliation["reconciled"]:
                val_cell.font = Font(name="Calibri", size=11, bold=True, color="2E7D32")
            else:
                val_cell.font = Font(name="Calibri", size=11, bold=True, color="C62828")
                
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 35
    
    # 2. Sheet 2: Transactions detailed ledger
    ws_txs = wb.create_sheet(title="Transactions")
    ws_txs.views.sheetView[0].showGridLines = True
    
    headers_txs = ["Date", "Description", "Amount", "Calculated Balance", "Category", "Review Status", "Source File", "Bank"]
    ws_txs.append(headers_txs)
    ws_txs.row_dimensions[1].height = 24
    
    for c_idx in range(1, 9):
        cell = ws_txs.cell(1, c_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for tx in transactions:
        # Sign withdrawals as negative, deposits as positive
        deb = tx.get("debit") if tx.get("debit") else 0.0
        cred = tx.get("credit") if tx.get("credit") else 0.0
        signed_amt = cred - deb
        
        # Check review status
        is_review = "Clean"
        if not tx["date"] or signed_amt == 0.0 or tx.get("balance") is None:
            is_review = "Needs Review"
            
        row_val = [
            tx["date"],
            tx["description"],
            signed_amt,
            tx["balance"] if tx["balance"] is not None else "",
            tx.get("category", ""),
            is_review,
            filename,
            bank_name
        ]
        ws_txs.append(row_val)
        
    # Format transactions cells
    for r_idx in range(2, ws_txs.max_row + 1):
        ws_txs.row_dimensions[r_idx].height = 18
        # Date alignment
        ws_txs.cell(r_idx, 1).alignment = Alignment(horizontal="center")
        # Amount formatting
        ws_txs.cell(r_idx, 3).number_format = "$#,##0.00"
        ws_txs.cell(r_idx, 4).number_format = "$#,##0.00"
        
        # Border
        for col_idx in range(1, 9):
            ws_txs.cell(r_idx, col_idx).border = thin_border
            
    # Auto-fit columns
    for col in ws_txs.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_txs.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Save to buffer
    wb.save(output)
    return output.getvalue()

def generate_annual_workbook(df_sorted, opening_balance):
    """
    Builds a consolidated annual Excel workbook from a sorted DataFrame.
    Inserts a completely blank separator row between transactions of different months.
    """
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    
    # 1. Summary Dashboard
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    title_font = Font(name="Calibri", size=16, bold=True, color="1F497D")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    accent_fill = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # Recalculate totals
    total_debits = pd.to_numeric(df_sorted['debit'], errors='coerce').fillna(0).sum()
    total_credits = pd.to_numeric(df_sorted['credit'], errors='coerce').fillna(0).sum()
    closing_bal = opening_balance
    if not df_sorted.empty:
        # Find last non-empty balance
        valid_bals = df_sorted['balance'].dropna()
        if not valid_bals.empty:
            closing_bal = float(valid_bals.iloc[-1])
            
    calculated_closing = opening_balance - total_debits + total_credits
    diff = round(closing_bal - calculated_closing, 2)
    reconciled = abs(diff) <= 0.05
    
    ws_summary.append(["ANNUAL ACCOUNTING CONSOLIDATED REPORT"])
    ws_summary.cell(1, 1).font = title_font
    ws_summary.row_dimensions[1].height = 30
    ws_summary.append([]) # spacer
    
    metrics = [
        ("Parameter", "Value"),
        ("Reporting Period", "Annual Workbook"),
        ("Transaction Count", len(df_sorted)),
        ("Opening Balance", opening_balance),
        ("Total Withdrawals (Debits)", -total_debits),
        ("Total Deposits (Credits)", total_credits),
        ("Closing Balance", closing_bal),
        ("Reconciliation Difference", diff),
        ("Reconciliation Status", "RECONCILED" if reconciled else "MISMATCH WARNING")
    ]
    
    for label, val in metrics:
        ws_summary.append([label, val])
        
    # Style Summary Grid
    ws_summary.merge_cells("A1:B1")
    ws_summary.row_dimensions[3].height = 20
    ws_summary.cell(3, 1).font = header_font
    ws_summary.cell(3, 1).fill = header_fill
    ws_summary.cell(3, 2).font = header_font
    ws_summary.cell(3, 2).fill = header_fill
    
    for r_idx in range(4, 13):
        ws_summary.cell(r_idx, 1).font = bold_font
        ws_summary.cell(r_idx, 1).fill = accent_fill
        ws_summary.cell(r_idx, 1).border = thin_border
        
        val_cell = ws_summary.cell(r_idx, 2)
        val_cell.font = normal_font
        val_cell.border = thin_border
        
        # Number formats
        if r_idx in (6, 7, 8, 9, 10):
            val_cell.number_format = "$#,##0.00"
            
        if r_idx == 11: # status warning coloring
            if reconciled:
                val_cell.font = Font(name="Calibri", size=11, bold=True, color="2E7D32")
            else:
                val_cell.font = Font(name="Calibri", size=11, bold=True, color="C62828")
                
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 35
    
    # 2. Sheet 2: Transactions detailed ledger
    ws_txs = wb.create_sheet(title="Transactions")
    ws_txs.views.sheetView[0].showGridLines = True
    
    headers_txs = ["Date", "Description", "Amount", "Calculated Balance", "Category", "Review Status", "Source File", "Bank"]
    ws_txs.append(headers_txs)
    ws_txs.row_dimensions[1].height = 24
    
    for c_idx in range(1, 9):
        cell = ws_txs.cell(1, c_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    prev_month_str = None
    
    for _, r in df_sorted.iterrows():
        # Check month transition to insert blank separator row
        date_str = str(r.get("date", ""))
        current_month_str = None
        if len(date_str) >= 7: # YYYY-MM
            current_month_str = date_str[:7]
            
        if prev_month_str is not None and current_month_str != prev_month_str:
            # Insert completely blank separator row
            ws_txs.append([""] * 8)
            # Make the blank row shorter and borderless
            ws_txs.row_dimensions[ws_txs.max_row].height = 10
            
        prev_month_str = current_month_str
        
        # Sign withdrawals as negative, deposits as positive
        deb = float(r["debit"]) if pd.notna(r["debit"]) and r["debit"] != "" else 0.0
        cred = float(r["credit"]) if pd.notna(r["credit"]) and r["credit"] != "" else 0.0
        signed_amt = cred - deb
        
        is_review = "Clean"
        if not r["date"] or signed_amt == 0.0 or pd.isna(r["balance"]) or r["balance"] == "":
            is_review = "Needs Review"
            
        row_val = [
            r["date"],
            r["description"],
            signed_amt,
            float(r["balance"]) if pd.notna(r["balance"]) and r["balance"] != "" else "",
            r.get("category", ""),
            is_review,
            r.get("source_file", ""),
            r.get("institution", "")
        ]
        ws_txs.append(row_val)
        
    # Style and format cell values
    for r_idx in range(2, ws_txs.max_row + 1):
        # Skip blank separator rows
        is_row_blank = all(ws_txs.cell(r_idx, col_idx).value == "" or ws_txs.cell(r_idx, col_idx).value is None for col_idx in range(1, 9))
        if is_row_blank:
            continue
            
        ws_txs.row_dimensions[r_idx].height = 18
        ws_txs.cell(r_idx, 1).alignment = Alignment(horizontal="center")
        
        ws_txs.cell(r_idx, 3).number_format = "$#,##0.00"
        ws_txs.cell(r_idx, 4).number_format = "$#,##0.00"
        
        for col_idx in range(1, 9):
            ws_txs.cell(r_idx, col_idx).border = thin_border
            
    # Auto-fit columns
    for col in ws_txs.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_txs.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    wb.save(output)
    return output.getvalue()
